"""Excel report generator — .xlsx with summary and detail sheets."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..engine.hasher import DuplicateGroup

_METHOD_LABELS = {
    "exact": "精准匹配",
    "perceptual": "感知哈希",
    "feature": "特征匹配",
    "video": "视频查重",
    "semantic": "AI语义",
}

_HEADER_FILL = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")


def _human_size(size: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024:
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{size:.1f} TB"


class ExcelReportGenerator:
    """Generate an Excel (.xlsx) duplicate report."""

    def generate(
        self,
        groups: list[DuplicateGroup],
        output_path: str | Path,
    ) -> Path:
        output_path = Path(output_path)
        wb = Workbook()

        self._write_summary(wb, groups)
        self._write_details(wb, groups)

        wb.save(str(output_path))
        return output_path.resolve()

    # ── sheets ───────────────────────────────────────────────

    def _write_summary(self, wb: Workbook, groups: list[DuplicateGroup]) -> None:
        ws = wb.active
        ws.title = "概览"

        headers = ["检测方法", "重复组数", "涉及文件数"]
        self._write_header_row(ws, headers)

        method_counts: dict[str, list[int]] = defaultdict(lambda: [0, 0])
        for g in groups:
            mc = method_counts[g.detection_method]
            mc[0] += 1
            mc[1] += len(g.files)

        row = 2
        for method in ("exact", "perceptual", "feature", "video", "semantic"):
            if method not in method_counts:
                continue
            cnt, files = method_counts[method]
            ws.cell(row=row, column=1, value=_METHOD_LABELS.get(method, method))
            ws.cell(row=row, column=2, value=cnt)
            ws.cell(row=row, column=3, value=files)
            row += 1

        self._auto_width(ws)

    def _write_details(self, wb: Workbook, groups: list[DuplicateGroup]) -> None:
        ws = wb.create_sheet("详细结果")

        headers = ["组ID", "检测方法", "相似度", "文件名", "文件路径", "文件大小", "尺寸(宽x高)"]
        self._write_header_row(ws, headers)

        row = 2
        for idx, g in enumerate(groups):
            fill = _GREEN_FILL if idx % 2 == 0 else None
            label = _METHOD_LABELS.get(g.detection_method, g.detection_method)
            similarity = f"{g.similarity_score * 100:.1f}%"

            for f in g.files:
                ws.cell(row=row, column=1, value=g.group_id)
                ws.cell(row=row, column=2, value=label)
                ws.cell(row=row, column=3, value=similarity)
                ws.cell(row=row, column=4, value=Path(f.file_path).name)
                ws.cell(row=row, column=5, value=f.file_path)
                ws.cell(row=row, column=6, value=_human_size(f.file_size))
                ws.cell(row=row, column=7, value=f"{f.width}x{f.height}")
                if fill:
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).fill = fill
                row += 1

        self._auto_width(ws)

    # ── helpers ───────────────────────────────────────────────

    @staticmethod
    def _write_header_row(ws, headers: list[str]) -> None:
        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _auto_width(ws) -> None:
        for col_cells in ws.columns:
            max_len = 0
            for cell in col_cells:
                try:
                    length = len(str(cell.value or ""))
                    if length > max_len:
                        max_len = length
                except Exception:
                    pass
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
